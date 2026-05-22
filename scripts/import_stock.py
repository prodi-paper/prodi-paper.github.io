#!/usr/bin/env python3
"""
Import/sync Prodiconseil stock from xlsx files → Supabase.
Strategy: full replacement (DELETE all → INSERT fresh) per run.
Skips: FABRICATION files, summary/aggregated files (0R, 0S, EXTERIEURS...).
"""

import os, re, json, time
import urllib.request, urllib.parse
import openpyxl

# ── Config ───────────────────────────────────────────────────────────────────
SUPABASE_URL = 'https://bvcgpdoukhcatjibmvnb.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ2Y2dwZG91a2hjYXRqaWJtdm5iIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzIyNzg5MjgsImV4cCI6MjA4Nzg1NDkyOH0.Ip3ykSUS9sajTH04yXBerOG1haBKMD1kAvMQNjnGL1Q'
IMAGE_BASE    = 'https://stock.prodi.net/albums/photo/{num}.jpg'

XLSX_DIRS = [
    '/Users/tantan/Desktop/dossier sans titre',
]

# Files/patterns to skip
SKIP_PATTERNS = [
    r'FABRICATION',
    r'^0[RS]?\s*-',          # 0R - STOCK DEPOT..., 0S - ..., 0 - EXTERIEURS...
    r'REUNIS',
    r'EXTERIEURS',
    r'EZPass',
    r'MIXTE',                 # aggregated files
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def sb_request(method, path, data=None, params=None, extra_headers=None):
    url = SUPABASE_URL + path
    if params:
        url += '?' + urllib.parse.urlencode(params)
    body = json.dumps(data).encode() if data is not None else None
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal',
    }
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read()
            return resp.status, json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()


def parse_price(val):
    """'760Eur/T' or 760.0 or '760' → 760.0"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val else None
    m = re.search(r'[\d]+(?:[.,]\d+)?', str(val))
    return float(m.group().replace(',', '.')) if m else None


def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def make_image_url(ref):
    if not ref:
        return None
    m = re.search(r'Photo_(\d+)', str(ref), re.IGNORECASE)
    if m:
        return IMAGE_BASE.format(num=m.group(1))
    return None


def detect_format(filename):
    fn = filename.upper()
    if 'BOBINE' in fn or fn.startswith('R') or 'REEL' in fn:
        return 'Bobine'
    if 'PALETTE' in fn or 'FORMAT' in fn or 'SHEET' in fn:
        return 'Palette'
    # Heuristic from code prefix
    base = os.path.basename(filename)
    m = re.match(r'\d+([RS])\s*-', base)
    if m:
        return 'Bobine' if m.group(1) == 'R' else 'Palette'
    return 'Bobine'


HEADER_KEYWORDS = {'qualit', 'quality', 'gr', 'gsm', 'laize', 'poids', 'couleur', 'color'}

def find_header_row(ws):
    """Find the row index (0-based) where the column headers are."""
    best_i, best_cells, best_score = None, None, 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row is None:
            continue
        cells = [str(c).strip().lower() if c is not None else '' for c in row]
        has_qualit = any('qualit' in c or c in ('quality','qualite','qualité') for c in cells)
        matches = sum(1 for c in cells if any(k in c for k in HEADER_KEYWORDS))
        # Must have quality keyword + at least 2 other matches to be a real header
        score = matches if has_qualit else 0
        if score > best_score:
            best_score, best_i, best_cells = score, i, cells
    if best_score >= 3:
        return best_i, best_cells
    return None, None


def parse_xlsx(filepath):
    """Parse one xlsx file → list of record dicts."""
    fmt = detect_format(os.path.basename(filepath))
    records = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        hdr_idx, headers = find_header_row(ws)
        if hdr_idx is None:
            print(f'  [SKIP] no header found: {os.path.basename(filepath)}')
            wb.close()
            return []

        # Map columns by keyword
        col = {}
        for i, h in enumerate(headers):
            # Use exact/priority matching to avoid 'qualit' matching 'ref qualite'
            if h in ('qualité','qualite','quality','qualit') and 'ref' not in h:
                col['quality'] = i
            elif h in ('gr','gsm') or (h == 'grammage'):
                col['gsm'] = i
            elif h in ('laize','width (mm)','width','largeur'):
                col['width'] = i
            elif 'diamètre' in h or 'diametre' in h or 'diameter' in h:
                col['diam'] = i
            elif h in ('mandrin','noyau','core (mm)','core'):
                col['noyau'] = i
            elif ('poids' in h or 'weight' in h) and 'ref' not in h:
                col['weight'] = i
            elif h in ('couleur','color','colour') or (('couleur' in h or 'color' in h) and 'ref' not in h):
                col['color'] = i
            elif h in ('details','détails','detail','détail'):
                col['details'] = i
            elif 'ref qualite' in h or ('ref' in h and 'qualit' in h):
                col['usine_ref'] = i   # "USINE 287" factory number
            elif 'emplacement' in h or 'location' in h:
                col['emplacement'] = i  # "OUR WAREHOUSE" / "FRANCE" depot
            elif 'depart usine' in h or ('usine' in h and 'ref' not in h and 'cout' not in h and 'qualit' not in h):
                if 'price' not in col: col['price'] = i
            elif 'cout' in h and 'fret' in h:
                col['price_cf'] = i
            elif h in ('longueur','length','long (mm)') or ('longueur' in h and 'ref' not in h):
                col['longueur'] = i
            elif ('photo' in h or 'picture' in h or 'view' in h) and 'ref' not in h:
                col['ref'] = i
            elif h == 'to view picture click here' or h == 'tpicture click here' or h == 'clic here':
                col['ref'] = i

        if 'quality' not in col:
            print(f'  [SKIP] no quality column: {os.path.basename(filepath)}')
            wb.close()
            return []

        all_rows = list(ws.iter_rows(values_only=True))
        data_rows = all_rows[hdr_idx + 1:]

        for row in data_rows:
            if not row or all(c is None for c in row):
                continue

            def g(key):
                idx = col.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            quality = clean_str(g('quality'))
            if not quality or quality.lower() in ('qualité', 'quality', 'total', 'totaux', 'sous-total'):
                continue
            # Skip obviously non-product rows
            if re.match(r'^(total|sous|nb\b|quantit)', quality.lower()):
                continue

            ref_raw = clean_str(g('ref'))
            # Extract the numeric ID or FABxxxx from Photo_ refs (strip stacked Photo_ prefixes)
            ref = None
            is_fab_ref = False
            if ref_raw:
                cleaned = re.sub(r'(?i)(?:photo_)+', '', str(ref_raw)).strip()
                if cleaned:
                    ref = 'Photo_' + cleaned
                    is_fab_ref = 'FAB' in cleaned.upper()

            weight_raw = g('weight')
            weight = None
            if isinstance(weight_raw, (int, float)) and weight_raw > 0:
                weight = float(weight_raw)
            elif weight_raw:
                weight = parse_price(weight_raw)

            gsm_raw = g('gsm')
            gsm = None
            if isinstance(gsm_raw, (int, float)):
                gsm = int(gsm_raw) if gsm_raw else None
            elif gsm_raw:
                try: gsm = int(float(str(gsm_raw).replace(',','.')))
                except: pass

            width_raw = g('width')
            width = None
            if isinstance(width_raw, (int, float)):
                width = int(width_raw) if width_raw else None
            elif width_raw:
                try: width = int(float(str(width_raw).replace(',','.')))
                except: pass

            noyau_raw = g('noyau')
            noyau = None
            if isinstance(noyau_raw, (int, float)):
                noyau = int(noyau_raw) if noyau_raw else None
            elif noyau_raw:
                try: noyau = int(float(str(noyau_raw).replace(',','.')))
                except: pass

            longueur_raw = g('longueur') or g('diam')  # diam used as longueur for palettes
            longueur = None
            if fmt == 'Palette' and longueur_raw:
                if isinstance(longueur_raw, (int, float)):
                    longueur = int(longueur_raw) if longueur_raw else None
                else:
                    try: longueur = int(float(str(longueur_raw).replace(',','.')))
                    except: pass

            price = parse_price(g('price'))

            emplacement_raw = g('emplacement')
            emplacement = clean_str(emplacement_raw) if emplacement_raw else None

            usine_raw = clean_str(g('usine_ref'))
            usine = None
            if usine_raw:
                # Normalisation alignée avec import_stock_ci.py::extract_usine.
                # Gère 'USINE U 171', 'USINE USDAM 171', 'REF 102', '007', etc.
                # Rejette 'USINE', 'USINE REPR', 'PAL.DIVERS...' (None).
                m = re.search(r'\bUSINE\s*[A-Z]*\s*(\d+)', usine_raw, re.IGNORECASE)
                if m:
                    usine = str(int(m.group(1)))
                else:
                    stripped = re.sub(r'^REF\s*', '', usine_raw, flags=re.IGNORECASE).strip()
                    usine = str(int(stripped)) if re.fullmatch(r'\d+', stripped) else None

            rec = {
                'quality':     quality,
                'color':       clean_str(g('color')),
                'details':     clean_str(g('details')),
                'gsm':         gsm,
                'width':       width,
                'weight':      weight,
                'price':       price,
                'ref':         ref,
                'noyau':       noyau,
                'format':      fmt,
                'image_url':   make_image_url(ref),
                'longueur':    longueur,
                'emplacement': emplacement,
                'usine':       usine,
            }
            # Remove None values
            rec = {k: v for k, v in rec.items() if v is not None}
            records.append(rec)

        wb.close()
    except Exception as e:
        print(f'  [ERROR] {os.path.basename(filepath)}: {e}')
    return records


def should_skip(filename):
    base = os.path.basename(filename)
    for pat in SKIP_PATTERNS:
        if re.search(pat, base, re.IGNORECASE):
            return True
    return False


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # 1. Collect all xlsx files
    xlsx_files = []
    for d in XLSX_DIRS:
        for fn in sorted(os.listdir(d)):
            if fn.endswith('.xlsx'):
                fp = os.path.join(d, fn)
                if should_skip(fn):
                    print(f'[SKIP] {fn}')
                else:
                    xlsx_files.append(fp)

    print(f'\n→ {len(xlsx_files)} files to process\n')

    # 2. Parse all files
    all_records = []
    file_counts = {}
    for fp in xlsx_files:
        bn = os.path.basename(fp)
        recs = parse_xlsx(fp)
        file_counts[bn] = len(recs)
        all_records.extend(recs)
        print(f'  {bn}: {len(recs)} records')

    print(f'\n→ Total parsed: {len(all_records)} records')

    # 3. Detect duplicates (same ref)
    ref_seen = {}
    duplicates = []
    deduped = []
    for rec in all_records:
        ref = rec.get('ref')
        if ref:
            if ref in ref_seen:
                duplicates.append({'ref': ref, 'quality': rec.get('quality'), 'first_quality': ref_seen[ref]})
            else:
                ref_seen[ref] = rec.get('quality')
                deduped.append(rec)
        else:
            deduped.append(rec)  # no ref → always include

    print(f'→ Duplicates detected (same Photo ref): {len(duplicates)}')
    if duplicates:
        for d in duplicates[:20]:
            print(f'   {d["ref"]} — quality={d["quality"]} (first seen: {d["first_quality"]})')
        if len(duplicates) > 20:
            print(f'   ... and {len(duplicates)-20} more')

    print(f'→ Records after dedup: {len(deduped)}\n')

    # 4. Get existing records from DB (paginated)
    print('Fetching existing DB records...')
    existing = []
    offset = 0
    PAGE = 1000
    while True:
        status, page = sb_request('GET', '/rest/v1/products',
            params={'select': 'id,ref', 'limit': PAGE, 'offset': offset},
            extra_headers={'Range-Unit': 'items', 'Range': f'{offset}-{offset+PAGE-1}'})
        if not isinstance(page, list):
            print(f'[ERROR] Failed to fetch page at offset {offset}: {status} {page}')
            break
        existing.extend(page)
        if len(page) < PAGE:
            break
        offset += PAGE
    print(f'→ Fetched {len(existing)} existing records')

    existing_refs = {r['ref']: r['id'] for r in existing if r.get('ref')}
    print(f'→ {len(existing)} records in DB ({len(existing_refs)} with refs)\n')

    # 5. Normalize all records to same keys (required for batch insert)
    ALL_KEYS = ['quality','color','details','gsm','width','weight','price','ref','noyau','format','image_url','longueur','emplacement','usine']
    def normalize(rec):
        return {k: rec.get(k) for k in ALL_KEYS}
    deduped = [normalize(r) for r in deduped]

    # Split into INSERT vs UPDATE
    to_insert = []
    to_update = []  # list of (id, record)
    for rec in deduped:
        ref = rec.get('ref')
        if ref and ref in existing_refs:
            to_update.append((existing_refs[ref], rec))
        else:
            to_insert.append(rec)

    # Records in DB no longer in any xlsx (sold/removed)
    new_refs = {r.get('ref') for r in deduped if r.get('ref')}
    to_delete_ids = [r['id'] for r in existing if r.get('ref') and r['ref'] not in new_refs]

    print(f'→ INSERT: {len(to_insert)} new records')
    print(f'→ UPDATE: {len(to_update)} existing records')
    print(f'→ DELETE: {len(to_delete_ids)} records no longer in stock\n')

    proceed = 'yes'
    if proceed != 'yes':
        print('Aborted.')
        return
    print('→ Confirmed.')

    # 6. Delete stale records
    if to_delete_ids:
        print(f'Deleting {len(to_delete_ids)} stale records...')
        CHUNK = 100
        deleted = 0
        for i in range(0, len(to_delete_ids), CHUNK):
            chunk = to_delete_ids[i:i+CHUNK]
            ids_str = ','.join(str(x) for x in chunk)
            status, _ = sb_request('DELETE', f'/rest/v1/products?id=in.({ids_str})')
            if status in (200, 204):
                deleted += len(chunk)
            else:
                print(f'  [ERROR] delete chunk: {status}')
        print(f'  Deleted {deleted} records')

    # 7. Batch INSERT
    if to_insert:
        print(f'Inserting {len(to_insert)} new records...')
        CHUNK = 200
        inserted = 0
        errors = 0
        for i in range(0, len(to_insert), CHUNK):
            chunk = to_insert[i:i+CHUNK]
            status, resp = sb_request('POST', '/rest/v1/products', data=chunk)
            if status in (200, 201):
                inserted += len(chunk)
                print(f'  Inserted {inserted}/{len(to_insert)}...')
            else:
                errors += 1
                print(f'  [ERROR] insert chunk {i}: {status} {str(resp)[:200]}')
            time.sleep(0.1)
        print(f'  → Inserted {inserted}, errors: {errors}')

    # 8. Batch UPDATE (PATCH by id)
    if to_update:
        print(f'Updating {len(to_update)} records...')
        updated = 0
        errors = 0
        CHUNK = 50
        for i in range(0, len(to_update), CHUNK):
            chunk = to_update[i:i+CHUNK]
            for rec_id, rec in chunk:
                status, resp = sb_request('PATCH', f'/rest/v1/products?id=eq.{rec_id}', data=rec)
                if status in (200, 204):
                    updated += 1
                else:
                    errors += 1
                    print(f'  [ERROR] update id={rec_id}: {status} {str(resp)[:100]}')
            print(f'  Updated {min(i+CHUNK, len(to_update))}/{len(to_update)}...')
            time.sleep(0.05)
        print(f'  → Updated {updated}, errors: {errors}')

    # 9. Summary
    print('\n━━━ DONE ━━━')
    status, final = sb_request('GET', '/rest/v1/products', params={'select': 'id', 'limit': 1},
        extra_headers={'Prefer': 'count=exact'})
    print(f'DB now has records (check manually on Supabase dashboard)')
    print(f'Duplicates detected across xlsx files: {len(duplicates)}')

if __name__ == '__main__':
    main()
