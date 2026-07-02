#!/usr/bin/env python3
"""
Import automatique du stock Prodiconseil depuis le mail quotidien.
- Se connecte à Gmail via IMAP
- Récupère le dernier mail de info@prodi.com (sujet "STOCK DÉTAILLÉ")
- Parse tous les fichiers Excel attachés
- Met à jour la table products dans Supabase

Usage:
  python3 import_stock_auto.py          # import depuis le dernier mail
  python3 import_stock_auto.py --dry    # dry run (affiche sans modifier la base)
"""

import imaplib, email, email.header, os, sys, json, re, tempfile, subprocess, urllib.request
from datetime import datetime
from collections import defaultdict

# ── CONFIG ──
IMAP_HOST = "imap.gmail.com"
IMAP_USER = os.environ.get("IMAP_USER", "eelbilia@gmail.com")
IMAP_PASS = os.environ["IMAP_PASS"]
SENDER = "info@prodi.com"

SUPABASE_URL = "https://bvcgpdoukhcatjibmvnb.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ2Y2dwZG91a2hjYXRqaWJtdm5iIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzIyNzg5MjgsImV4cCI6MjA4Nzg1NDkyOH0.Ip3ykSUS9sajTH04yXBerOG1haBKMD1kAvMQNjnGL1Q"
MGMT_TOKEN = os.environ["SUPABASE_MGMT_TOKEN"]

ALL_KEYS = ['quality','color','details','gsm','width','longueur','noyau','weight','price','ref','usine','emplacement','zone','format','image_url','source']

DRY_RUN = '--dry' in sys.argv

# ── HELPERS ──
def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def parse_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    m = re.match(r'(\d+(?:\.\d+)?)', s.replace(',', '.'))
    return float(m.group(1)) if m else None

def extract_usine(ref_str):
    """Normalise les valeurs Sage 'REF QUALITÉ' en code usine pur (str(int)).
    Gère : '287' / '007' / 'USINE 287' / 'USINE U 171' / 'USINE USDAM 171' /
    'USINE USP 159' / 'USINE USNT 3' / 'REF 102'. Retourne None pour les
    valeurs non-numériques ('USINE', 'USINE REPR', 'PAL.DIVERS...').
    """
    if ref_str is None: return None
    s = str(ref_str).strip()
    if not s or 'eur' in s.lower(): return None
    m = re.search(r'\bUSINE\s*[A-Z]*\s*(\d+)', s, re.IGNORECASE)
    if m: return str(int(m.group(1)))
    stripped = re.sub(r'^REF\s*', '', s, flags=re.IGNORECASE).strip()
    if re.fullmatch(r'\d+', stripped): return str(int(stripped))
    return None

# Sépare la colonne EMPLACEMENT/LOCATION (col 12 des fichiers par qualité) en
# (emplacement, zone). Le mail "AVEC ZONE" met l'ALLÉE précise dans cette colonne
# (ex "6KD", "3UG, 3VG") à la place de "OUR WAREHOUSE" :
#   - allée détectée → emplacement="OUR WAREHOUSE" (garde le filtre NOTRE DÉPÔT)
#                      + zone=<allée du jour>
#   - sinon (DIRECT USINE, FAB DEPART ST OUEN, FRANCE…) → emplacement tel quel, zone=None.
_AISLE_RE = re.compile(r'^\s*\d{1,3}[A-Z]{1,3}(?:\s*,\s*\d{1,3}[A-Z]{1,3})*\s*$', re.IGNORECASE)
def split_location(val):
    if val is None:
        return None, None
    s = str(val).strip()
    if not s:
        return None, None
    if _AISLE_RE.match(s):
        return "OUR WAREHOUSE", s.upper()
    return s, None

# ── STEP 1: FETCH EMAIL ──
def fetch_latest_stock_email():
    log("Connexion IMAP...")
    mail = imaplib.IMAP4_SSL(IMAP_HOST, 993)
    mail.login(IMAP_USER, IMAP_PASS)
    mail.select('INBOX')

    status, msgs = mail.search(None, f'(FROM "{SENDER}")')
    msg_ids = msgs[0].split()
    if not msg_ids:
        log("Aucun mail de stock trouvé")
        mail.logout()
        return None, None

    # Chaque matin, info@prodi.com envoie 2 mails (~15s d'écart) :
    # "STOCK DÉTAILLÉ AVEC ZONE" puis "STOCK DÉTAILLÉ" (sans zone). On veut la
    # version AVEC ZONE. On NE peut PAS se fier à l'ordre d'arrivée (le sans-zone
    # arrive en dernier) ni prendre aveuglément le dernier mail (pourrait être un
    # autre courrier). On choisit par SUJET, du plus récent au plus ancien.
    def _subject(mid):
        _, d = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
        raw = email.message_from_bytes(d[0][1]).get('Subject', '')
        return ''.join(
            (b.decode(enc or 'utf-8', 'replace') if isinstance(b, bytes) else b)
            for b, enc in email.header.decode_header(raw)
        ).upper()

    recent = list(reversed(msg_ids))[:12]
    subj = {mid: _subject(mid) for mid in recent}
    latest_id = next((mid for mid in recent if 'STOCK' in subj[mid] and 'ZONE' in subj[mid]), None)
    if latest_id is None:  # repli : un "STOCK DÉTAILLÉ" récent (sans zone)
        latest_id = next((mid for mid in recent if 'STOCK' in subj[mid]), None)
    if latest_id is None:  # ultime repli : comportement historique
        latest_id = msg_ids[-1]

    status, data = mail.fetch(latest_id, '(RFC822)')
    msg = email.message_from_bytes(data[0][1])

    date_str = msg.get('Date', '')
    log(f"Mail choisi: {msg.get('Subject', '?')} — {date_str}")

    # Extract attachments to temp dir
    tmpdir = tempfile.mkdtemp(prefix='prodi_stock_')
    attachments = []
    for part in msg.walk():
        if part.get_content_disposition() == 'attachment':
            fname = part.get_filename()
            if fname and fname.endswith('.xlsx'):
                fpath = os.path.join(tmpdir, fname)
                with open(fpath, 'wb') as f:
                    f.write(part.get_payload(decode=True))
                attachments.append(fpath)

    log(f"Pièces jointes Excel: {len(attachments)}")
    mail.logout()
    return tmpdir, attachments

# ── STEP 2: PARSE EXCEL FILES ──
def parse_all_files(files):
    import openpyxl

    all_products = []

    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=False, data_only=True)
        except:
            continue
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            wb.close()
            continue

        first_vals = [str(v).strip() if v else '' for v in rows_raw[0]]

        if 'Référence' in first_vals:
            for row in rows_raw[1:]:
                if row[0] is None and row[1] is None: continue
                quality = str(row[1]).strip() if row[1] else None
                if not quality: continue
                ref_val = str(row[0]).strip() if row[0] else None
                empl, zone = split_location(row[14] if len(row) > 14 else None)
                all_products.append({
                    'ref': f"Photo_{ref_val}" if ref_val else None,
                    'quality': quality,
                    'color': str(row[2]).strip() if row[2] else None,
                    'details': str(row[3]).strip() if row[3] else None,
                    'gsm': int(row[5]) if row[5] and row[5] != 0 else None,
                    'width': int(row[6]) if row[6] else None,
                    'longueur': int(row[7]) if row[7] and row[7] != 0 else None,
                    'noyau': int(row[9]) if row[9] and row[9] != 0 else None,
                    'weight': float(row[10]) if row[10] else None,
                    'price': parse_price(row[11]),
                    'usine': extract_usine(str(row[13]).strip() if row[13] else None),
                    'emplacement': empl,
                    'zone': zone,
                    'format': 'Palette' if quality.startswith('S') else 'Bobine',
                })
        else:
            header_idx = header_row = None
            for i, row in enumerate(rows_raw):
                vals = [str(v).strip().lower() if v else '' for v in row]
                if 'qualité' in vals:
                    header_idx, header_row = i, row
                    break
            if header_idx is None:
                wb.close()
                continue

            col_map = {}
            for ci, cv in enumerate(header_row):
                if cv is None: continue
                s = str(cv).strip().lower()
                if 'qualit' in s and 'ref' not in s: col_map['quality'] = ci
                elif 'couleur' in s: col_map['color'] = ci
                elif 'detail' in s: col_map['details'] = ci
                elif s.startswith('gr') and len(s) <= 3: col_map['gsm'] = ci
                elif 'laize' in s: col_map['width'] = ci
                elif 'longueur' in s: col_map['longueur'] = ci
                elif 'diam' in s: col_map['longueur'] = ci
                elif 'mandrin' in s: col_map['noyau'] = ci
                elif 'poids' in s: col_map['weight'] = ci
                elif 'depart' in s or 'exwork' in s: col_map['price'] = ci
                elif 'ref' in s and 'qualit' in s: col_map['usine'] = ci
                elif 'emplacement' in s: col_map['emplacement'] = ci

            def g(row, key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row): return None
                return row[idx]

            for row in rows_raw[header_idx+2:]:
                if len(row) < 5: continue
                quality = g(row, 'quality')
                if quality: quality = str(quality).strip()
                if not quality: continue
                ref_photo = str(row[0]).strip() if row[0] else None
                gsm_val = g(row,'gsm'); width_val = g(row,'width'); long_val = g(row,'longueur')
                noyau_val = g(row,'noyau'); weight_val = g(row,'weight')
                empl, zone = split_location(g(row,'emplacement'))
                all_products.append({
                    'ref': ref_photo,
                    'quality': quality,
                    'color': str(g(row,'color')).strip() if g(row,'color') else None,
                    'details': str(g(row,'details')).strip() if g(row,'details') else None,
                    'gsm': int(gsm_val) if gsm_val and isinstance(gsm_val,(int,float)) else None,
                    'width': int(width_val) if width_val and isinstance(width_val,(int,float)) else None,
                    'longueur': int(long_val) if long_val and isinstance(long_val,(int,float)) and long_val != 0 else None,
                    'noyau': int(noyau_val) if noyau_val and isinstance(noyau_val,(int,float)) and noyau_val != 0 else None,
                    'weight': float(weight_val) if weight_val and isinstance(weight_val,(int,float)) else None,
                    'price': parse_price(g(row,'price')),
                    'usine': extract_usine(str(g(row,'usine')).strip() if g(row,'usine') else None),
                    'emplacement': empl,
                    'zone': zone,
                    'format': 'Bobine' if quality.startswith('R') else 'Palette',
                })
        wb.close()

    # Fix refs
    for p in all_products:
        if p['ref']:
            while 'Photo_Photo_' in p['ref']: p['ref'] = p['ref'].replace('Photo_Photo_','Photo_')
            if not p['ref'].startswith('Photo_'): p['ref'] = 'Photo_' + p['ref']

    all_products = [p for p in all_products if p['quality'] not in ('GRIS / GREY',)]
    all_products = [p for p in all_products if not (p['color'] and p['color'].replace('.','').isdigit())]

    # Dedup by ref
    by_ref = defaultdict(list)
    for p in all_products: by_ref[p['ref']].append(p)
    deduped = [max(variants, key=lambda p: sum(1 for v in p.values() if v is not None)) for variants in by_ref.values()]

    # Extract image URLs from hyperlinks
    log("Extraction des URLs photos...")
    ref_to_url = {}
    for fp in files:
        try:
            import openpyxl as oxl
            wb2 = oxl.load_workbook(fp, read_only=False, data_only=False)
        except: continue
        ws2 = wb2.active
        for row in ws2.iter_rows(min_row=1):
            cell = row[0]
            if cell.hyperlink and cell.hyperlink.target and 'stock.prodi.net' in str(cell.hyperlink.target):
                ref_val = str(cell.value).strip() if cell.value else None
                if ref_val and ref_val != 'To view picture click here':
                    key = f"Photo_{ref_val}"
                    while 'Photo_Photo_' in key: key = key.replace('Photo_Photo_','Photo_')
                    ref_to_url[key] = cell.hyperlink.target
        wb2.close()

    img_count = 0
    for p in deduped:
        if p['ref'] in ref_to_url:
            p['image_url'] = ref_to_url[p['ref']]; img_count += 1
        else:
            p['image_url'] = None

    # Uniform keys
    for p in deduped:
        # 'email' = stock du jour (vs 'inventaire' = complément DOV statique).
        # Le catalogue public filtre source=neq.inventaire → un NULL ici serait
        # exclu aussi (sémantique PostgREST des NULL) : toujours expliciter.
        p['source'] = 'email'
        for k in ALL_KEYS:
            if k not in p: p[k] = None
        for k in list(p.keys()):
            if k not in ALL_KEYS: del p[k]

    w = sum(1 for p in deduped if p['weight'] is not None)
    log(f"Produits: {len(deduped)}, avec poids: {w}, avec photo: {img_count}")
    return deduped

# ── STEP 3: UPDATE SUPABASE ──
def update_supabase(products):
    if DRY_RUN:
        log("DRY RUN — aucune modification en base")
        return

    log("Suppression des anciens produits...")
    subprocess.run(['curl','-s','-o','/dev/null','-X','DELETE',
        f'{SUPABASE_URL}/rest/v1/products?id=gt.0',
        '-H',f'apikey: {ANON_KEY}','-H',f'Authorization: Bearer {ANON_KEY}',
        '-H','Prefer: return=minimal'], capture_output=True)

    BATCH = 500
    total = len(products)
    success = errors = 0
    for i in range(0, total, BATCH):
        batch = products[i:i+BATCH]
        tmpfile = '/tmp/prodi_batch.json'
        with open(tmpfile,'w') as f: json.dump(batch, f, ensure_ascii=False)
        result = subprocess.run(['curl','-s','-w','%{http_code}','-X','POST',
            f'{SUPABASE_URL}/rest/v1/products',
            '-H',f'apikey: {ANON_KEY}','-H',f'Authorization: Bearer {ANON_KEY}',
            '-H','Content-Type: application/json','-H','Prefer: return=minimal',
            '-d','@/tmp/prodi_batch.json'], capture_output=True, text=True)
        code = result.stdout[-3:]
        if code == '201':
            success += len(batch)
        else:
            errors += len(batch)
            log(f"  ERREUR batch {i//BATCH+1}: {result.stdout[:-3][:200]}")

    log(f"Insertion: {success} OK, {errors} erreurs")

    # Zones (allées) : désormais fournies FRAÎCHES par l'email "AVEC ZONE"
    # (col EMPLACEMENT → split_location → champ zone, inséré directement). On
    # n'applique donc PLUS le fichier statique correction_zone.xlsx (périmé).
    APPLY_STATIC_ZONES = False
    zone_candidates = [
        "/Users/tantan/Desktop/dossier sans titre 2/correction_zone.xlsx",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "correction_zone.xlsx"),
    ]
    zone_file = next((p for p in zone_candidates if os.path.exists(p)), None)
    if APPLY_STATIC_ZONES and zone_file:
        log("Application des zones/allées...")
        import openpyxl
        wb = openpyxl.load_workbook(zone_file, read_only=True, data_only=True)
        ws = wb.active
        ref_zone = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ref, famille, zone, corr_zone, c_zone, coul = row
            if not ref: continue
            ref_str = str(ref).strip()
            # FIX (2026-05-29) : Quand C_ZONE == 'OK', la zone validée est dans
            # col 2 (ZONES), pas col 3 (CORRECTIONS_ZONE) — col 3 contient juste
            # le flag "OK". Quand C_ZONE == 'FAUX', col 3 propose des corrections.
            if c_zone == 'OK' and zone:
                ref_zone[ref_str] = str(zone).strip()
            elif c_zone == 'FAUX' and corr_zone:
                ref_zone[ref_str] = str(corr_zone).strip()
            elif zone:
                ref_zone[ref_str] = str(zone).strip()
        wb.close()

        # Apply zones via REST API PATCH (anon key works thanks to RLS)
        applied = 0
        failed = 0
        for ref, zone in ref_zone.items():
            photo_ref = f"Photo_{ref}"
            url = f'{SUPABASE_URL}/rest/v1/products?ref=eq.{photo_ref}'
            payload = json.dumps({"zone": zone}).encode()
            req = urllib.request.Request(url, data=payload, method='PATCH', headers={
                'apikey': ANON_KEY,
                'Authorization': f'Bearer {ANON_KEY}',
                'Content-Type': 'application/json',
                'Prefer': 'return=minimal',
            })
            try:
                with urllib.request.urlopen(req) as resp:
                    if resp.status in (200, 204):
                        applied += 1
                    else:
                        failed += 1
            except Exception:
                failed += 1
        log(f"Zones appliquées: {applied} OK, {failed} erreurs (sur {len(ref_zone)} refs)")

# ── STEP 4: COMPLÉMENT INVENTAIRE ──
# L'export DOV complet (scripts/inventaire_complement.json, ~9 200 réfs, généré
# par convert_inventaire_toutarticle.py) couvre TOUT l'ERP alors que le mail
# quotidien n'en donne que ~5 900 : sans lui, les scans d'inventaire des réfs
# manquantes finissent « à vérifier ». On insère ici les réfs absentes du stock
# email du jour, avec source='inventaire' (le catalogue B2B public les filtre,
# l'app inventaire les voit). À refaire chaque matin : update_supabase() vide
# toute la table avant de réinsérer. GARDER SYNCHRO avec import_stock_ci.py.
def insert_inventory_complement(email_products):
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inventaire_complement.json')
    if not os.path.exists(path):
        log("Complément inventaire absent — étape sautée")
        return
    with open(path) as f:
        complement = json.load(f)
    # Réfs email préfixées Photo_ ; complément en réf ERP brute.
    strip = lambda r: r[6:] if r.startswith('Photo_') else r
    existing = {strip(str(p['ref'])) for p in email_products if p.get('ref')}
    to_insert = [p for p in complement if strip(str(p['ref'])) not in existing]
    log(f"Complément inventaire: {len(complement)} réfs DOV, {len(to_insert)} absentes du stock email")
    if DRY_RUN:
        log("DRY RUN — complément non inséré")
        return
    BATCH = 500
    success = errors = 0
    for i in range(0, len(to_insert), BATCH):
        batch = to_insert[i:i+BATCH]
        tmpfile = '/tmp/prodi_batch.json'
        with open(tmpfile,'w') as f: json.dump(batch, f, ensure_ascii=False)
        result = subprocess.run(['curl','-s','-w','%{http_code}','-X','POST',
            f'{SUPABASE_URL}/rest/v1/products',
            '-H',f'apikey: {ANON_KEY}','-H',f'Authorization: Bearer {ANON_KEY}',
            '-H','Content-Type: application/json','-H','Prefer: return=minimal',
            '-d','@/tmp/prodi_batch.json'], capture_output=True, text=True)
        if result.stdout[-3:] == '201':
            success += len(batch)
        else:
            errors += len(batch)
            log(f"  ERREUR complément batch {i//BATCH+1}: {result.stdout[:-3][:200]}")
    log(f"Complément inventaire inséré: {success} OK, {errors} erreurs")

# ── STEP 5: RÉ-APPARIEMENT INVENTAIRE ──
# La FK inventaire_lignes.product_id est ON DELETE SET NULL et update_supabase()
# régénère tous les ids de products → chaque import détache TOUTES les lignes
# d'inventaire. Ré-appariement par référence via la RPC (migration 019 du repo
# prodi_arrivages). NB : la clé anon n'a PAS le droit d'exécuter cette RPC
# (grant service_role/authenticated) — cette version locale est best-effort,
# la voie normale est import_stock_ci.py (service_role). GARDER SYNCHRO.
def rematch_inventaire_lignes():
    if DRY_RUN:
        log("DRY RUN — ré-appariement inventaire sauté")
        return
    result = subprocess.run(['curl','-s','-X','POST',
        f'{SUPABASE_URL}/rest/v1/rpc/rematch_inventaire_product_ids',
        '-H',f'apikey: {ANON_KEY}','-H',f'Authorization: Bearer {ANON_KEY}',
        '-H','Content-Type: application/json','-d','{}'], capture_output=True, text=True)
    log(f"Ré-appariement inventaire (best-effort anon): {result.stdout.strip() or '?'}")

# ── MAIN ──
if __name__ == '__main__':
    log("=== Import stock Prodiconseil ===")
    tmpdir, files = fetch_latest_stock_email()
    if not files:
        log("Aucun fichier à traiter")
        sys.exit(0)

    products = parse_all_files(files)
    update_supabase(products)
    insert_inventory_complement(products)
    rematch_inventaire_lignes()

    # Cleanup
    import shutil
    shutil.rmtree(tmpdir, ignore_errors=True)

    log(f"=== Terminé ! {len(products)} produits importés ===")
