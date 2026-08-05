#!/usr/bin/env python3
"""
Import automatique du stock Prodiconseil depuis le mail quotidien.
- Se connecte à Gmail via IMAP
- Récupère le dernier mail de info@prodi.com (sujet "STOCK DÉTAILLÉ")
- Parse tous les fichiers Excel attachés
- Met à jour la table products dans Supabase

Usage:
  python3 import_stock_auto.py          # import depuis le dernier mail
  python3 import_stock_auto.py --dry    # dry run (affiche sans modifier la base)
"""

import imaplib, email, email.header, os, sys, json, re, tempfile, subprocess
from datetime import datetime
from collections import defaultdict

# ── CONFIG (from environment variables) ──
IMAP_HOST = "imap.gmail.com"
IMAP_USER = os.environ["IMAP_USER"]
IMAP_PASS = os.environ["IMAP_PASS"]
SENDER = "info@prodi.com"

SUPABASE_URL = os.environ["SUPABASE_URL"]
ANON_KEY = os.environ["SUPABASE_ANON_KEY"]
MGMT_TOKEN = os.environ["SUPABASE_MGMT_TOKEN"]
# service_role bypasses RLS — required for DELETE/INSERT since RLS hardening (2026-05-01)
SERVICE_ROLE = os.environ["SUPABASE_SERVICE_ROLE"]

ALL_KEYS = ['quality','color','details','gsm','width','longueur','noyau','weight','price','ref','usine','emplacement','zone','format','image_url','source','reserve_client','reserve_piece','promo','date_arrivee']

DRY_RUN = '--dry' in sys.argv

# ── HELPERS ──
def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def parse_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    m = re.match(r'(\d+(?:\.\d+)?)', s.replace(',', '.'))
    return float(m.group(1)) if m else None

def extract_usine(ref_str):
    """Normalise les valeurs Sage 'REF QUALITÉ' en code usine pur (str(int)).
    Gère : '287' / '007' / 'USINE 287' / 'USINE U 171' / 'USINE USDAM 171' /
    'USINE USP 159' / 'USINE USNT 3' / 'REF 102'. Retourne None pour les
    valeurs non-numériques ('USINE', 'USINE REPR', 'PAL.DIVERS...').
    """
    if ref_str is None: return None
    s = str(ref_str).strip()
    if not s or 'eur' in s.lower(): return None
    m = re.search(r'\bUSINE\s*[A-Z]*\s*(\d+)', s, re.IGNORECASE)
    if m: return str(int(m.group(1)))
    stripped = re.sub(r'^REF\s*', '', s, flags=re.IGNORECASE).strip()
    if re.fullmatch(r'\d+', stripped): return str(int(stripped))
    return None

# Sépare la colonne EMPLACEMENT/LOCATION (col 12 des fichiers par qualité) en
# (emplacement, zone). Le mail "AVEC ZONE" met l'ALLÉE précise dans cette colonne
# (ex "6KD", "11BG", "3UG, 3VG") à la place de "OUR WAREHOUSE" :
#   - allée détectée  → emplacement="OUR WAREHOUSE" (garde le filtre NOTRE DÉPÔT)
#                        + zone=<allée du jour>
#   - sinon (DIRECT USINE, FAB DEPART ST OUEN, FRANCE, REF 102…) → emplacement tel
#     quel, zone=None.
_AISLE_RE = re.compile(r'^\s*\d{1,3}[A-Z]{1,3}(?:\s*,\s*\d{1,3}[A-Z]{1,3})*\s*$', re.IGNORECASE)
def split_location(val):
    if val is None:
        return None, None
    s = str(val).strip()
    if not s:
        return None, None
    if _AISLE_RE.match(s):
        return "OUR WAREHOUSE", s.upper()
    return s, None

# ── STEP 1: FETCH EMAIL ──
def fetch_latest_stock_email():
    log("Connexion IMAP...")
    mail = imaplib.IMAP4_SSL(IMAP_HOST, 993)
    mail.login(IMAP_USER, IMAP_PASS)
    mail.select('INBOX')

    status, msgs = mail.search(None, f'(FROM "{SENDER}")')
    msg_ids = msgs[0].split()
    if not msg_ids:
        log("Aucun mail de stock trouvé")
        mail.logout()
        return None, None

    # Chaque matin, info@prodi.com envoie 2 mails (~15s d'écart) :
    # "STOCK DÉTAILLÉ AVEC ZONE" puis "STOCK DÉTAILLÉ" (sans zone). On veut la
    # version AVEC ZONE. On NE peut PAS se fier à l'ordre d'arrivée (le sans-zone
    # arrive en dernier) ni prendre aveuglément le dernier mail (pourrait être un
    # autre courrier). On choisit par SUJET, du plus récent au plus ancien.
    def _subject(mid):
        _, d = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
        raw = email.message_from_bytes(d[0][1]).get('Subject', '')
        return ''.join(
            (b.decode(enc or 'utf-8', 'replace') if isinstance(b, bytes) else b)
            for b, enc in email.header.decode_header(raw)
        ).upper()

    # Depuis le 2026-07-02 la source est le mail « STOCK COMPLET AVEC LES
    # RESERVATION » (PJ unique INV_toutarticle.xlsx = export DOV complet,
    # ~9 300 lignes, allées en colonne DP_CODE). L'ancien « STOCK DÉTAILLÉ
    # AVEC ZONE » (83 PJ par qualité) ne part plus depuis le 2026-07-01 ;
    # « STOCK DÉTAILLÉ » (sans zone) arrive encore mais n'est PAS utilisé.
    # PAS de repli vers un autre mail : mieux vaut échouer (alerte Resend)
    # que d'importer la mauvaise source et vider le catalogue.
    recent = list(reversed(msg_ids))[:12]
    subj = {mid: _subject(mid) for mid in recent}
    latest_id = next((mid for mid in recent if 'STOCK' in subj[mid] and ('COMPLET' in subj[mid] or 'RESERVATION' in subj[mid])), None)
    if latest_id is None:
        log("ERREUR: aucun mail « STOCK COMPLET AVEC LES RESERVATION » récent trouvé")
        mail.logout()
        return None, None

    status, data = mail.fetch(latest_id, '(RFC822)')
    msg = email.message_from_bytes(data[0][1])

    date_str = msg.get('Date', '')
    log(f"Mail choisi: {msg.get('Subject', '?')} — {date_str}")

    # Extract attachments to temp dir
    tmpdir = tempfile.mkdtemp(prefix='prodi_stock_')
    attachments = []
    for part in msg.walk():
        if part.get_content_disposition() == 'attachment':
            fname = part.get_filename()
            if fname and fname.endswith('.xlsx'):
                fpath = os.path.join(tmpdir, fname)
                with open(fpath, 'wb') as f:
                    f.write(part.get_payload(decode=True))
                attachments.append(fpath)

    log(f"Pièces jointes Excel: {len(attachments)}")
    mail.logout()
    return tmpdir, attachments

# ── STEP 2: PARSE DOV (INV_toutarticle.xlsx) ──
# Une seule PJ : feuille DOV_export, ~9 300 lignes, TOUT l'ERP. Deux destins :
#   - familles papier/carton vendables ET stock > 0 → source='email'
#     (visibles sur le catalogue B2B public)
#   - le reste (machines UMAC/UMAN, frais WFRA, fret WFRE, écarts ECART,
#     quantité nulle) → source='inventaire' (invisible du catalogue, mais
#     présent pour la reconnaissance des scans d'inventaire)
# Mapping vérifié colonne par colonne le 2026-07-02. Remplace l'ancien parsing
# des 83 PJ par qualité (mail « AVEC ZONE » disparu le 2026-07-01).
HIDDEN_FAMILIES = {'UMAC', 'UMAN', 'WFRA', 'WFRE', 'ECART'}

def parse_dov(files):
    import openpyxl
    _prix_aberrants = 0
    fp = next((f for f in files if 'toutarticle' in os.path.basename(f).lower().replace('_', '')), files[0] if files else None)
    if not fp:
        return []
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip().strip('"') if h else '' for h in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    if 'REF' not in idx or 'CODE_FAM' not in idx:
        log(f"ERREUR: structure inattendue ({os.path.basename(fp)}) — colonnes {header[:6]}")
        wb.close()
        return []

    def g(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    def clean(v):
        s = str(v).strip() if v is not None else ''
        return s if s and s != '-' else None

    def num(v):
        try:
            f = float(v)
            return f if f > 0 else None
        except (TypeError, ValueError):
            return None

    by_ref = {}
    for row in rows_iter:
        ref = clean(g(row, 'REF'))
        if not ref:
            continue
        qty = num(g(row, 'QTSTO')) or 0
        # Doublons de réf (multi-lots) : on garde la ligne à la plus grosse qté.
        if ref in by_ref and by_ref[ref][0] >= qty:
            continue
        fam_code = clean(g(row, 'CODE_FAM'))
        fam_lib = clean(g(row, 'FAM')) or ''
        visible = fam_code not in HIDDEN_FAMILIES and qty > 0

        dp = clean(g(row, 'DP_CODE'))
        zone = dp.upper() if dp and _AISLE_RE.match(dp) else None
        depot = clean(g(row, 'NOM_DEPOT'))
        emplacement = 'OUR WAREHOUSE' if depot == 'A-PRODI SAINT-OUEN' else depot
        # Bobines : le diamètre (HDIAM) alimente `longueur`, même convention
        # que l'ancien import par qualité (piège 'diam' du CLAUDE.md).
        longueur = num(g(row, 'LONG')) or num(g(row, 'HDIAM'))
        # details = désignation + champs structurés (chips de l'app inventaire).
        parts = []
        for col in ('AR_Langue1', 'DETAIL', 'FIBRE', 'BACK', 'FINITION', 'QUALITE', 'TEINTE'):
            v = clean(g(row, col))
            if v and v.upper() not in (p.upper() for p in parts):
                parts.append(v)
        # Le DOV n'a pas d'hyperliens photo : URLs déterministes stock.prodi.net
        # pour les réfs numériques (le site gère les 404 par fallback visuel).
        image_url = f"https://stock.prodi.net/albums/photo/{ref}.jpg" if ref.isdigit() else None

        # Règle PROMO (18/07/2026, Ethan) : toute réf numérique < 900000 est du
        # vieux stock re-listé — JAMAIS une vraie réservation (les CODE_CLI
        # qu'on y trouve sont des scories) → promo, prix -30 %, résa levée.
        # Date d'arrivée Sage (DATECREA) → products.date_arrivee (21/07/2026).
        _dc = g(row, 'DATECREA')
        try:
            date_arrivee = _dc.date().isoformat() if hasattr(_dc, 'date') else (str(_dc)[:10] if _dc and str(_dc)[:4].isdigit() else None)
        except Exception:
            date_arrivee = None

        promo = ref.isdigit() and int(ref) < 900000
        # Prix (04/08) : PUNET vaut le prix d'ACHAT à ±1 % sur 52 % des lignes
        # du DOV et est SOUS l'achat sur 20 % — c'est une valorisation de
        # stock, pas un prix de vente. On prend donc le MAX de PUNET et
        # AR_PRIXVEN (tarif article, rempli à ~100 %). Garde-fou PAR CANDIDAT :
        # famille papier (R*/S*) au-dessus de 3 €/kg = prix unitaire saisi
        # dans un champ €/kg → candidat écarté (évite les 731 600 €/T), sans
        # perdre l'autre valeur si elle est saine.
        cands = [num(g(row, 'PUNET')), num(g(row, 'AR_PRIXVEN'))]
        if fam_code[:1] in ('R', 'S'):
            if any(c > 3 for c in cands if c):
                _prix_aberrants += 1  # logs Actions PUBLICS : ne pas y imprimer réfs/prix
            cands = [c for c in cands if c and c <= 3]
        prix = max([c for c in cands if c], default=None)
        if promo and prix:
            prix = round(prix * 0.7, 4)
        by_ref[ref] = (qty, {
            'ref': f"Photo_{ref}",
            'quality': fam_code,
            'color': clean(g(row, 'COULEUR')),
            'details': ' '.join(parts) or None,
            'gsm': int(num(g(row, 'GRS')) or 0) or None,
            'width': int(num(g(row, 'LARG')) or 0) or None,
            'longueur': int(longueur) if longueur else None,
            'noyau': int(num(g(row, 'MANDRIN')) or 0) or None,
            'weight': num(g(row, 'PNET')),
            'price': prix,
            'usine': extract_usine(clean(g(row, 'EMPLACEMENT'))),
            'emplacement': emplacement,
            'zone': zone,
            'format': 'Bobine' if fam_lib.startswith('BOB') else 'Palette' if fam_lib.startswith('PAL') else None,
            'image_url': image_url,
            'source': 'email' if visible else 'inventaire',
            # Réservation Sage : code client + bon de préparation (BPxxxxx).
            # (QTRES existe dans le fichier mais reste à 0 — la réservation
            # s'exprime par CODE_CLI/CODE_PIECE.)
            'reserve_client': None if promo else (clean(g(row, 'CODE_CLI')) or None),
            'reserve_piece': None if promo else (clean(g(row, 'CODE_PIECE')) or None),
            'promo': promo,
            'date_arrivee': date_arrivee,
        })
    wb.close()

    products = [p for _, p in by_ref.values()]
    visibles = sum(1 for p in products if p['source'] == 'email')
    log(f"DOV: {len(products)} réfs uniques — {visibles} visibles catalogue, {len(products) - visibles} inventaire seul")
    if _prix_aberrants:
        log(f"prix aberrants ignorés : {_prix_aberrants} (valeurs non listées — logs publics)")
    return products

# ── STEP 3: UPDATE SUPABASE ──
def update_supabase(products):
    if DRY_RUN:
        log("DRY RUN — aucune modification en base")
        return

    log("Suppression des anciens produits...")
    subprocess.run(['curl','-s','-o','/dev/null','-X','DELETE',
        f'{SUPABASE_URL}/rest/v1/products?id=gt.0',
        '-H',f'apikey: {SERVICE_ROLE}','-H',f'Authorization: Bearer {SERVICE_ROLE}',
        '-H','Prefer: return=minimal'], capture_output=True)

    BATCH = 500
    total = len(products)
    success = errors = 0
    for i in range(0, total, BATCH):
        batch = products[i:i+BATCH]
        tmpfile = '/tmp/prodi_batch.json'
        with open(tmpfile,'w') as f: json.dump(batch, f, ensure_ascii=False)
        result = subprocess.run(['curl','-s','-w','%{http_code}','-X','POST',
            f'{SUPABASE_URL}/rest/v1/products',
            '-H',f'apikey: {SERVICE_ROLE}','-H',f'Authorization: Bearer {SERVICE_ROLE}',
            '-H','Content-Type: application/json','-H','Prefer: return=minimal',
            '-d','@/tmp/prodi_batch.json'], capture_output=True, text=True)
        code = result.stdout[-3:]
        if code == '201':
            success += len(batch)
        else:
            errors += len(batch)
            log(f"  ERREUR batch {i//BATCH+1}: {result.stdout[:-3][:200]}")

    log(f"Insertion: {success} OK, {errors} erreurs")

    # Zones (allées) : désormais fournies FRAÎCHES par l'email "AVEC ZONE"
    # (colonne EMPLACEMENT → split_location → champ zone, inséré directement).
    # On n'applique donc PLUS le fichier statique correction_zone.xlsx (figé au
    # 24/04) qui écraserait les allées du jour par des valeurs périmées.
    APPLY_STATIC_ZONES = False
    zone_file = os.path.join(os.path.dirname(__file__), "correction_zone.xlsx")
    if APPLY_STATIC_ZONES and os.path.exists(zone_file):
        log("Application des zones/allées...")
        import openpyxl
        wb = openpyxl.load_workbook(zone_file, read_only=True, data_only=True)
        ws = wb.active
        ref_zone = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ref, famille, zone, corr_zone, c_zone, coul = row
            if not ref: continue
            ref_str = str(ref).strip()
            # FIX (2026-05-29) : Quand C_ZONE == 'OK', la zone validée est dans
            # col 2 (ZONES), pas col 3 (CORRECTIONS_ZONE) — col 3 contient juste
            # le flag "OK" et finissait écrit en base à la place du vrai code.
            # Quand C_ZONE == 'FAUX', col 3 propose des corrections (souvent
            # multi-codes ex "7MG, 7NG, 7OG") : on les écrit telles quelles.
            if c_zone == 'OK' and zone:
                ref_zone[ref_str] = str(zone).strip()
            elif c_zone == 'FAUX' and corr_zone:
                ref_zone[ref_str] = str(corr_zone).strip()
            elif zone:
                ref_zone[ref_str] = str(zone).strip()
        wb.close()

        SQL_URL = "https://api.supabase.com/v1/projects/bvcgpdoukhcatjibmvnb/database/query"
        refs_list = list(ref_zone.items())
        for i in range(0, len(refs_list), 500):
            batch = refs_list[i:i+500]
            cases = []
            where_refs = []
            for ref, zone in batch:
                zone_escaped = zone.replace("'", "''")
                photo_ref = f"Photo_{ref}"
                cases.append(f"WHEN ref = '{photo_ref}' THEN '{zone_escaped}'")
                where_refs.append(f"'{photo_ref}'")
            sql = f"UPDATE products SET zone = CASE {' '.join(cases)} ELSE zone END WHERE ref IN ({','.join(where_refs)});"
            subprocess.run(['curl','-s','-X','POST', SQL_URL,
                '-H',f'Authorization: Bearer {MGMT_TOKEN}',
                '-H','Content-Type: application/json',
                '-d', json.dumps({"query": sql})], capture_output=True)
        log(f"Zones appliquées: {len(ref_zone)} refs")


# ── STEP 3bis: FILET IMAGE_URL (server-side, idempotent) ──
# Constaté le 2026-07-02 : sur le runner CI, ~1 000 lignes sont arrivées en
# base avec image_url NULL alors que le parse local du MÊME fichier avec le
# MÊME code produit bien l'URL (cause non identifiée : parse déterministe,
# types de cellules identiques, log d'insertion propre). Ce filet SQL
# synthétise l'URL manquante pour toute réf numérique, quel que soit
# l'environnement — le catalogue ne peut plus perdre ses photos.
def backfill_image_urls():
    if DRY_RUN:
        return
    sql = ("update products set image_url = 'https://stock.prodi.net/albums/photo/' "
           "|| substring(ref from 7) || '.jpg' "
           "where image_url is null and ref ~ '^Photo_[0-9]+$';")
    subprocess.run(['curl','-s','-X','POST',
        'https://api.supabase.com/v1/projects/bvcgpdoukhcatjibmvnb/database/query',
        '-H',f'Authorization: Bearer {MGMT_TOKEN}',
        '-H','Content-Type: application/json',
        '-d', json.dumps({'query': sql})], capture_output=True)
    log("Filet image_url appliqué (réfs numériques sans URL)")

# ── STEP 4: RÉ-APPARIEMENT INVENTAIRE ──
# La FK inventaire_lignes.product_id est ON DELETE SET NULL et update_supabase()
# régénère tous les ids de products → chaque import détache TOUTES les lignes
# d'inventaire (elles s'affichent « hors catalogue » dans l'app). On les
# ré-apparie par référence (stable) via la RPC rematch_inventaire_product_ids
# (migration 019 du repo prodi_arrivages).
def rematch_inventaire_lignes():
    if DRY_RUN:
        log("DRY RUN — ré-appariement inventaire sauté")
        return
    result = subprocess.run(['curl','-s','-X','POST',
        f'{SUPABASE_URL}/rest/v1/rpc/rematch_inventaire_product_ids',
        '-H',f'apikey: {SERVICE_ROLE}','-H',f'Authorization: Bearer {SERVICE_ROLE}',
        '-H','Content-Type: application/json','-d','{}'], capture_output=True, text=True)
    log(f"Ré-appariement inventaire: {result.stdout.strip() or '?'} lignes rattachées")

# ── MAIN ──
if __name__ == '__main__':
    log("=== Import stock Prodiconseil ===")
    tmpdir, files = fetch_latest_stock_email()
    if not files:
        log("ERREUR: mail STOCK COMPLET introuvable — import annulé (base intacte)")
        sys.exit(1)

    products = parse_dov(files)
    # Garde-fou : update_supabase VIDE la table avant d'insérer. Un fichier
    # anormalement petit (mauvaise PJ, structure changée) annulerait le
    # catalogue entier → on refuse et on laisse la base d'hier en place.
    if len(products) < 5000:
        log(f"ABANDON: {len(products)} produits parsés (< 5000) — base NON touchée")
        sys.exit(1)
    # Garde-fou PRIX (04/08) : si les colonnes PUNET/AR_PRIXVEN disparaissent
    # ou changent de nom dans le DOV, on publierait un catalogue sans prix.
    # Couverture normale ≈ 99 % — sous 60 %, on refuse (mail d'alerte Resend
    # via le step if:failure du workflow) et la base d'hier reste en place.
    _vis = [p for p in products if p['source'] == 'email']
    _avec_prix = sum(1 for p in _vis if p.get('price'))
    if _vis and _avec_prix / len(_vis) < 0.6:
        log(f"ABANDON: {_avec_prix}/{len(_vis)} produits visibles avec prix (< 60 %) — colonnes prix du DOV suspectes, base NON touchée")
        sys.exit(1)
    update_supabase(products)
    backfill_image_urls()
    rematch_inventaire_lignes()
    try:
        sync_offres_fab()
    except Exception as e:
        log(f"offres fab: échec NON bloquant — {e}")

    # Cleanup
    import shutil
    shutil.rmtree(tmpdir, ignore_errors=True)

    log(f"=== Terminé ! {len(products)} produits importés ===")


# ── STEP 5 (bonus) : OFFRES FABRICATION ──
def sync_offres_fab():
    """Publie les PJ « * FABRICATION.xlsx » du mail « STOCK DÉTAILLÉ » sur le
    bucket PUBLIC offres-fab + manifest.json — servies par le bouton
    « Fabrication » du catalogue (05/08). Les Excel téléchargés par les
    clients sont EXACTEMENT ceux de Sage, aucun retraitement. Échec = log
    seulement : l'import du stock ne doit jamais tomber pour ça."""
    import urllib.request, urllib.parse
    from datetime import date
    mail = imaplib.IMAP4_SSL(IMAP_HOST, 993)
    mail.login(IMAP_USER, IMAP_PASS)
    mail.select('INBOX')
    _, msgs = mail.search(None, f'(FROM "{SENDER}")')
    ids = list(reversed(msgs[0].split()))[:12]

    def _subj(mid):
        _, d = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT)])')
        raw = email.message_from_bytes(d[0][1]).get('Subject', '')
        return ''.join(
            (b.decode(enc or 'utf-8', 'replace') if isinstance(b, bytes) else b)
            for b, enc in email.header.decode_header(raw)
        ).upper()

    # « STOCK DÉTAILLÉ » (83 PJ par qualité) — TAILL évite l'accent encodé
    mid = next((m for m in ids if 'STOCK' in _subj(m) and 'TAILL' in _subj(m)), None)
    if mid is None:
        log("offres fab: mail STOCK DÉTAILLÉ introuvable — bucket inchangé")
        mail.logout()
        return
    _, data = mail.fetch(mid, '(RFC822)')
    msg = email.message_from_bytes(data[0][1])
    mail.logout()

    def _up(name, blob, ctype):
        req = urllib.request.Request(
            f"{SUPABASE_URL}/storage/v1/object/offres-fab/{urllib.parse.quote(name)}",
            data=blob, method='POST',
            headers={'Authorization': f'Bearer {SERVICE_ROLE}',
                     'Content-Type': ctype, 'x-upsert': 'true'})
        urllib.request.urlopen(req, timeout=60)

    import io as _io
    import zipfile as _zip
    import openpyxl as _px
    XCT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def _extrait_offres(blob):
        """Lignes d'offres du fichier (menu Usine déplié, 05/08) : mappées par
        les EN-TÊTES (GR, Laize, Longueur/Diamètre, Poids, DEPART, REF QUALITE)."""
        wb = _px.load_workbook(_io.BytesIO(blob), read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        hi = None
        ix = {}
        for i, r in enumerate(rows):
            vals = [str(v or '') for v in r]
            if 'GR' in vals and any('Qualit' in v for v in vals):
                hi = i
                for j, h in enumerate(vals):
                    hl = re.sub(r'\s+', ' ', h.lower())  # « REF\nQUALITE » a un saut de ligne
                    if h == 'GR':
                        ix['g'] = j
                    elif 'laize' in hl:
                        ix['laize'] = j
                    elif 'longueur' in hl or 'diam' in hl:
                        ix['long'] = j
                    elif 'poids' in hl:
                        ix['poids'] = j
                    elif 'depart' in hl:
                        ix['prix'] = j
                    elif 'ref qualite' in hl or 'ref qualité' in hl:
                        ix['usine'] = j
                break
        out = []
        if hi is None:
            return out

        def _n(v):
            try:
                return int(float(str(v).replace(' ', '').replace(',', '.')))
            except Exception:
                return None
        for r in rows[hi + 1:]:
            vals = [str(v or '') for v in r]
            if not any(v.startswith('Photo_') for v in vals):
                continue
            mp = re.search(r'(\d+(?:[.,]\d+)?)\s*Eur?/T', str(r[ix['prix']] if 'prix' in ix else ''), re.I)
            mu = re.search(r'USINE\s*0*(\d+)', str(r[ix['usine']] if 'usine' in ix else ''), re.I)
            out.append({'g': _n(r[ix['g']]) if 'g' in ix else None,
                        'laize': _n(r[ix['laize']]) if 'laize' in ix else None,
                        'long': _n(r[ix['long']]) if 'long' in ix else None,
                        'poids': _n(r[ix['poids']]) if 'poids' in ix else None,
                        'prix': int(float(mp.group(1))) if mp else None,
                        'usine': mu.group(1) if mu else None})
        out.sort(key=lambda o: (o['g'] or 0, o['laize'] or 0))
        return out

    def _usines_rows(blob):
        """({usine: [n° de lignes]}, nb de réfs Photo_) — un fichier SANS réf
        est un gabarit vide (vécu SADH 05/08) : on l'exclut du menu."""
        wb = _px.load_workbook(_io.BytesIO(blob), read_only=True, data_only=True)
        out = {}
        refs = 0
        for row in wb.active.iter_rows():
            for c in row:
                sv = str(c.value or '')
                if sv.startswith('Photo_'):
                    refs += 1
                mu = re.match(r'.*USINE\s*0*(\d+)', sv)
                if mu:
                    out.setdefault(mu.group(1), []).append(c.row)
        wb.close()
        return out, refs

    def _variante(blob, hide_rows):
        """Copie du zip d'origine avec hidden=1 sur les lignes des autres
        usines (feuille 1) — logos/styles byte-identiques (openpyxl PERDRAIT
        les images en resauvant, d'où la chirurgie XML)."""
        zin = _zip.ZipFile(_io.BytesIO(blob))
        buf = _io.BytesIO()
        with _zip.ZipFile(buf, 'w', _zip.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/worksheets/sheet1.xml':
                    x = data.decode('utf-8')
                    for rn in hide_rows:
                        x = re.sub(f'<row r="{rn}"(?![0-9])', f'<row r="{rn}" hidden="1"', x, count=1)
                    data = x.encode('utf-8')
                zout.writestr(item, data)
        return buf.getvalue()

    manifest = []
    for part in msg.walk():
        if part.get_content_disposition() != 'attachment':
            continue
        fname = ''.join(
            (b.decode(enc or 'utf-8', 'replace') if isinstance(b, bytes) else b)
            for b, enc in email.header.decode_header(part.get_filename() or '')
        )
        if 'FABRICATION' not in fname.upper() or not fname.lower().endswith('.xlsx'):
            continue
        m = re.match(r'\d+([RS]) - (.+?) FABRICATION', fname)
        code = m.group(2).strip() if m else fname
        forme = 'Bobine' if (m and m.group(1) == 'R') else 'Format'
        safe = fname.replace(' ', '_')
        blob = part.get_payload(decode=True)
        # Variantes PAR USINE (choix d'usine dans le menu Fabrication, 05/08)
        variants = []
        try:
            ur, _refs = _usines_rows(blob)
            if _refs == 0:
                log(f"offres fab: {fname} VIDE (0 réf) — exclu du menu")
                continue
            _up(safe, blob, XCT)
            allrows = set(r for rows in ur.values() for r in rows)
            for us, rows in sorted(ur.items(), key=lambda kv: -len(kv[1])):
                vname = safe.replace('.xlsx', f'__USINE_{us}.xlsx')
                _up(vname, _variante(blob, sorted(allrows - set(rows))), XCT)
                variants.append({'usine': us, 'fichier': vname})
        except Exception as e:
            log(f"offres fab: variantes usines KO pour {fname} — {e}")
            try:
                _up(safe, blob, XCT)  # au moins le fichier complet
            except Exception:
                pass
        manifest.append({'fichier': safe, 'nom': fname, 'code': code, 'forme': forme, 'usines': variants, 'nb': _refs, 'offres': _extrait_offres(blob)})
    if manifest:
        _up('manifest.json',
            json.dumps({'date': date.today().isoformat(), 'fichiers': manifest},
                       ensure_ascii=False).encode(), 'application/json')
    log(f"offres fab: {len(manifest)} Excel + {sum(len(x['usines']) for x in manifest)} variantes usine publiés")
